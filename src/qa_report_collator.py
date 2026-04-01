"""
Step 12 — QA report collator.
Runs multi-agent QA verification on the pipeline outputs and writes:
  /output/qa/qa_report.json
  /output/qa/qa_findings.jsonl
  /output/qa/qa_disagreements.jsonl
  /output/qa/qa_human_review.xlsx

QA agents:
  - qa-recall-auditor-01/02    : check for over-rejection (missed SaaS words)
  - qa-noise-auditor-01/02     : check for under-rejection (noise slipped in)
  - qa-semantic-auditor-01/02  : check semantic label quality
  - qa-output-auditor-01/02    : check output file integrity
  - qa-chief-reviewer          : final QA verdict

QA uses the same pipeline entry point (runs on a sample of the real output).
"""

import datetime
import json
import os
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import anthropic

from config import (
    CLAUDE_MODEL,
    MAX_RESPONSE_TOKENS,
    OUT_QA_DISAGREEMENTS,
    OUT_QA_FINDINGS,
    OUT_QA_HUMAN_REVIEW_XLSX,
    OUT_QA_REPORT,
    OUT_REJECTED_WORDS,
    OUT_RUN_SUMMARY,
    OUT_SAAS_WORDS,
    PIPELINE_VERSION,
    API_RETRY_ATTEMPTS,
    API_RETRY_BASE_DELAY,
)
from utils import (
    append_jsonl,
    extract_json,
    get_logger,
    read_jsonl,
    read_json,
    with_retry,
    write_json,
    write_jsonl,
)

log = get_logger("qa_report_collator")

QA_SAMPLE_SIZE = int(os.environ.get("QA_SAMPLE_SIZE", "100"))

# ---------------------------------------------------------------------------
# QA Agent prompts
# ---------------------------------------------------------------------------

QA_RECALL_AUDITOR_PROMPT = """\
You are qa-recall-auditor.
Review a sample of ACCEPTED and REJECTED words from the SaaS word extraction pipeline.
Identify any REJECTED words that appear to have genuine SaaS title potential
(false negatives / over-rejection).

For each concerning rejection, provide a finding with severity (low/medium/high).
"""

QA_NOISE_AUDITOR_PROMPT = """\
You are qa-noise-auditor.
Review a sample of ACCEPTED words from the SaaS word extraction pipeline.
Identify any ACCEPTED words that appear to be noise, non-words, or clearly
unsuitable for SaaS titles (false positives / over-acceptance).

For each concerning acceptance, provide a finding with severity (low/medium/high).
"""

QA_SEMANTIC_AUDITOR_PROMPT = """\
You are qa-semantic-auditor.
Review a sample of ACCEPTED words and their assigned labels
(functional / brandable / ambiguous).
Identify any words where the label appears incorrect.

For each mislabeled word, provide a finding with the correct suggested label.
"""

QA_OUTPUT_AUDITOR_PROMPT = """\
You are qa-output-auditor.
You are given the run_summary statistics from the SaaS word extraction pipeline.
Identify any suspicious patterns such as:
- Unexpectedly high or low acceptance rate
- Unusual label distribution
- Missing fields or schema issues

Provide findings with severity (low/medium/high) and recommendations.
"""

QA_CHIEF_PROMPT = """\
You are qa-chief-reviewer.
You receive all findings from the recall, noise, semantic, and output QA auditors.
Your job is to:
1. Summarise the overall QA result (pass/pass_with_warnings/fail)
2. List the top 5 most critical findings
3. Provide actionable recommendations

Be thorough but fair. The pipeline prioritises recall over precision by design.
"""


# ---------------------------------------------------------------------------
# QA runner
# ---------------------------------------------------------------------------

class QACollator:
    def __init__(self):
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            raise EnvironmentError("ANTHROPIC_API_KEY not set")
        self.client = anthropic.Anthropic(api_key=api_key)

    def _call(self, system: str, user: str, agent_id: str) -> dict:
        def _attempt():
            response = self.client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=MAX_RESPONSE_TOKENS,
                system=system,
                messages=[{"role": "user", "content": user}],
            )
            return extract_json(response.content[0].text)

        return with_retry(_attempt, attempts=API_RETRY_ATTEMPTS,
                          base_delay=API_RETRY_BASE_DELAY, logger=log)

    def _sample_records(self) -> tuple[list[dict], list[dict]]:
        """Load and sample from final output JSONL files."""
        saas = read_jsonl(OUT_SAAS_WORDS) if OUT_SAAS_WORDS.exists() else []
        rejected = read_jsonl(OUT_REJECTED_WORDS) if OUT_REJECTED_WORDS.exists() else []

        # Stratified sample: general + borderline (risk_flags) + noise candidates
        borderline = [r for r in saas if r.get("risk_flags")]
        general_accepted = [r for r in saas if not r.get("risk_flags")]
        sample_size_each = max(10, QA_SAMPLE_SIZE // 4)

        sampled_accepted = (
            random.sample(borderline, min(len(borderline), sample_size_each)) +
            random.sample(general_accepted, min(len(general_accepted), sample_size_each * 2))
        )
        sampled_rejected = random.sample(rejected, min(len(rejected), sample_size_each))

        log.info("QA sample: %d accepted, %d rejected",
                 len(sampled_accepted), len(sampled_rejected))
        return sampled_accepted, sampled_rejected

    def _run_recall_audit(
        self, accepted: list[dict], rejected: list[dict]
    ) -> tuple[list[dict], list[dict]]:
        """Two recall auditors review the sample."""
        findings_all: list[dict] = []

        sample_text = (
            "ACCEPTED (sample):\n" +
            "\n".join(f"- {r['normalized_word']} [{r.get('primary_label','?')}] "
                      f"risk={r.get('risk_flags',[])}  why={r.get('why_accept',[])[:2]}"
                      for r in accepted[:50]) +
            "\n\nREJECTED (sample):\n" +
            "\n".join(f"- {r['normalized_word']}  reason={r.get('reject_reason','?')}"
                      for r in rejected[:50])
        )

        for auditor_id in ["qa-recall-auditor-01", "qa-recall-auditor-02"]:
            user_msg = (
                f"Review this sample and identify any over-rejections.\n\n"
                f"{sample_text}\n\n"
                f"Response format:\n"
                f'{{"findings": [{{"word": "...", "issue": "over_rejected", '
                f'"argument": "...", "severity": "low|medium|high"}}]}}'
            )
            try:
                result = self._call(QA_RECALL_AUDITOR_PROMPT, user_msg, auditor_id)
                for f in result.get("findings", []):
                    f["auditor"] = auditor_id
                    f["audit_type"] = "recall"
                    findings_all.append(f)
            except Exception as exc:
                log.error("%s failed: %s", auditor_id, exc)

        disagreements = [f for f in findings_all if f.get("severity") in ("medium", "high")]
        return findings_all, disagreements

    def _run_noise_audit(self, accepted: list[dict]) -> tuple[list[dict], list[dict]]:
        """Two noise auditors review accepted words."""
        findings_all: list[dict] = []

        sample_text = "\n".join(
            f"- {r['normalized_word']} [{r.get('primary_label','?')}]  "
            f"confidence={r.get('confidence',0):.2f}  flags={r.get('risk_flags',[])}"
            for r in accepted[:80]
        )

        for auditor_id in ["qa-noise-auditor-01", "qa-noise-auditor-02"]:
            user_msg = (
                f"Review these accepted words and identify any that should NOT have been accepted.\n\n"
                f"{sample_text}\n\n"
                f"Response format:\n"
                f'{{"findings": [{{"word": "...", "issue": "over_accepted", '
                f'"argument": "...", "severity": "low|medium|high"}}]}}'
            )
            try:
                result = self._call(QA_NOISE_AUDITOR_PROMPT, user_msg, auditor_id)
                for f in result.get("findings", []):
                    f["auditor"] = auditor_id
                    f["audit_type"] = "noise"
                    findings_all.append(f)
            except Exception as exc:
                log.error("%s failed: %s", auditor_id, exc)

        disagreements = [f for f in findings_all if f.get("severity") in ("medium", "high")]
        return findings_all, disagreements

    def _run_semantic_audit(self, accepted: list[dict]) -> tuple[list[dict], list[dict]]:
        """Two semantic auditors check label quality."""
        findings_all: list[dict] = []

        sample_text = "\n".join(
            f"- {r['normalized_word']} → label={r.get('primary_label','?')}"
            for r in accepted[:80]
        )

        for auditor_id in ["qa-semantic-auditor-01", "qa-semantic-auditor-02"]:
            user_msg = (
                f"Review these label assignments. functional=SaaS verb/feature noun, "
                f"brandable=product brand word, ambiguous=unclear.\n\n"
                f"{sample_text}\n\n"
                f"Response format:\n"
                f'{{"findings": [{{"word": "...", "issue": "wrong_label", '
                f'"current_label": "...", "suggested_label": "...", '
                f'"argument": "...", "severity": "low|medium|high"}}]}}'
            )
            try:
                result = self._call(QA_SEMANTIC_AUDITOR_PROMPT, user_msg, auditor_id)
                for f in result.get("findings", []):
                    f["auditor"] = auditor_id
                    f["audit_type"] = "semantic"
                    findings_all.append(f)
            except Exception as exc:
                log.error("%s failed: %s", auditor_id, exc)

        disagreements = [f for f in findings_all if f.get("severity") in ("medium", "high")]
        return findings_all, disagreements

    def _run_output_audit(self) -> tuple[list[dict], list[dict]]:
        """Two output auditors check file integrity and stats."""
        findings_all: list[dict] = []
        summary = read_json(OUT_RUN_SUMMARY) if OUT_RUN_SUMMARY.exists() else {}

        summary_text = json.dumps(summary, indent=2)

        for auditor_id in ["qa-output-auditor-01", "qa-output-auditor-02"]:
            user_msg = (
                f"Review this pipeline run summary for anomalies.\n\n"
                f"{summary_text}\n\n"
                f"Response format:\n"
                f'{{"findings": [{{"issue": "...", "detail": "...", '
                f'"severity": "low|medium|high"}}]}}'
            )
            try:
                result = self._call(QA_OUTPUT_AUDITOR_PROMPT, user_msg, auditor_id)
                for f in result.get("findings", []):
                    f["auditor"] = auditor_id
                    f["audit_type"] = "output"
                    findings_all.append(f)
            except Exception as exc:
                log.error("%s failed: %s", auditor_id, exc)

        disagreements = [f for f in findings_all if f.get("severity") in ("medium", "high")]
        return findings_all, disagreements

    def _run_chief_review(self, all_findings: list[dict]) -> dict:
        """QA chief reviews all findings and produces final verdict."""
        findings_text = json.dumps(all_findings[:100], indent=2)

        user_msg = (
            f"Review all QA findings and produce a final QA verdict.\n\n"
            f"All findings:\n{findings_text}\n\n"
            f"Response format:\n"
            f'{{"qa_verdict": "pass|pass_with_warnings|fail", '
            f'"top_findings": ["...", "..."], '
            f'"recommendations": ["..."], '
            f'"critical_count": 0, "warning_count": 0, "info_count": 0}}'
        )
        try:
            return self._call(QA_CHIEF_PROMPT, user_msg, "qa-chief-reviewer")
        except Exception as exc:
            log.error("qa-chief-reviewer failed: %s", exc)
            return {
                "qa_verdict": "pass_with_warnings",
                "top_findings": ["Chief reviewer unavailable"],
                "recommendations": ["Manual review recommended"],
                "critical_count": 0, "warning_count": 1, "info_count": 0,
            }

    def run(self) -> None:
        """Run full QA pipeline and write all QA output files."""
        log.info("Starting QA collation …")

        accepted_sample, rejected_sample = self._sample_records()

        recall_findings, recall_disagreements = self._run_recall_audit(
            accepted_sample, rejected_sample)
        noise_findings, noise_disagreements = self._run_noise_audit(accepted_sample)
        semantic_findings, semantic_disagreements = self._run_semantic_audit(accepted_sample)
        output_findings, output_disagreements = self._run_output_audit()

        all_findings = (
            recall_findings + noise_findings +
            semantic_findings + output_findings
        )
        all_disagreements = (
            recall_disagreements + noise_disagreements +
            semantic_disagreements + output_disagreements
        )

        chief_result = self._run_chief_review(all_findings)

        # Write QA output files
        write_jsonl(OUT_QA_FINDINGS, all_findings)
        write_jsonl(OUT_QA_DISAGREEMENTS, all_disagreements)

        qa_report = {
            "pipeline_version": PIPELINE_VERSION,
            "qa_timestamp": datetime.datetime.utcnow().isoformat() + "Z",
            "sample_accepted": len(accepted_sample),
            "sample_rejected": len(rejected_sample),
            "total_findings": len(all_findings),
            "total_disagreements": len(all_disagreements),
            "finding_breakdown": {
                "recall": len(recall_findings),
                "noise": len(noise_findings),
                "semantic": len(semantic_findings),
                "output": len(output_findings),
            },
            **chief_result,
        }
        write_json(OUT_QA_REPORT, qa_report)
        log.info("QA verdict: %s  findings=%d  disagreements=%d",
                 chief_result.get("qa_verdict"), len(all_findings), len(all_disagreements))

        _write_qa_human_review(all_findings, qa_report)
        log.info("QA collation complete.")


# ---------------------------------------------------------------------------
# QA human review XLSX
# ---------------------------------------------------------------------------

def _write_qa_human_review(findings: list[dict], report: dict) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: all findings
    ws_f = wb.active
    ws_f.title = "qa_findings"
    cols = ["auditor", "audit_type", "word", "issue", "argument",
            "severity", "current_label", "suggested_label", "detail"]
    ws_f.append(cols)
    for f in findings:
        ws_f.append([f.get(c, "") for c in cols])

    # Header styling
    _HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HEADER_FONT = Font(color="FFFFFF", bold=True)
    for cell in ws_f[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws_f.freeze_panes = "A2"
    ws_f.auto_filter.ref = ws_f.dimensions

    # Sheet 2: QA report summary
    ws_r = wb.create_sheet("qa_report")
    ws_r.append(["Key", "Value"])
    for cell in ws_r[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for k, v in report.items():
        ws_r.append([k, json.dumps(v) if isinstance(v, (list, dict)) else v])

    OUT_QA_HUMAN_REVIEW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_QA_HUMAN_REVIEW_XLSX)
    log.info("Wrote %s", OUT_QA_HUMAN_REVIEW_XLSX)


# ---------------------------------------------------------------------------
# Module entry point
# ---------------------------------------------------------------------------

def run() -> None:
    """Step 12: run QA and write all QA outputs."""
    collator = QACollator()
    collator.run()
