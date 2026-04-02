"""
Step 10 — Human review exporter.
Generates XLSX and CSV files from the same source records as the JSONL output.
All files are derived from the same origin data to prevent content divergence.

Output:
  /output/human_review/saas_words_review.xlsx
    Sheets: accepted_words, borderline_words, rejected_words, summary, qa_findings
  /output/human_review/saas_words_review.csv   (accepted + borderline only)
  /output/human_review/rejected_words_review.xlsx
"""

import csv
import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import (
    OUT_REJECTED_WORDS,
    OUT_REJECTED_REVIEW_XLSX,
    OUT_SAAS_WORDS,
    OUT_SAAS_REVIEW_CSV,
    OUT_SAAS_REVIEW_XLSX,
    PIPELINE_VERSION,
)
from utils import get_logger, iter_jsonl

log = get_logger("human_review_exporter")

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

ACCEPTED_COLS = [
    "word", "normalized_word", "decision", "primary_label", "candidate_modes",
    "confidence", "consensus_support", "consensus_oppose", "consensus_abstain",
    "why_accept", "risk_flags", "review_priority", "source_file", "source_line",
    "pipeline_version", "manual_note", "manual_override",
]

REJECTED_COLS = [
    "word", "normalized_word", "decision", "reject_reason",
    "consensus_support", "consensus_oppose", "consensus_abstain",
    "source_file", "source_line", "pipeline_version", "manual_note",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BORDER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RISK_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _list_to_str(v) -> str:
    if isinstance(v, list):
        return "; ".join(str(i) for i in v)
    return str(v) if v is not None else ""


def _review_priority(rec: dict) -> str:
    """Assign review priority based on risk flags and confidence."""
    flags = rec.get("risk_flags", [])
    confidence = rec.get("confidence", 1.0)
    if flags or confidence < 0.6:
        return "high"
    if confidence < 0.75:
        return "medium"
    return "low"


def _flatten_record(rec: dict, cols: list[str]) -> list:
    """Flatten a record dict to a row matching `cols`."""
    consensus = rec.get("consensus", {})
    row = []
    for col in cols:
        if col == "consensus_support":
            row.append(consensus.get("support", ""))
        elif col == "consensus_oppose":
            row.append(consensus.get("oppose", ""))
        elif col == "consensus_abstain":
            row.append(consensus.get("abstain", ""))
        elif col == "review_priority":
            row.append(_review_priority(rec))
        elif col in ("candidate_modes", "why_accept", "risk_flags", "reject_reason"):
            row.append(_list_to_str(rec.get(col, [])))
        elif col in ("manual_note", "manual_override"):
            row.append("")  # empty for human to fill
        else:
            row.append(rec.get(col, ""))
    return row


def _style_sheet(ws, cols: list[str]):
    """Apply header style, freeze top row, and set column widths."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths (heuristic)
    width_map = {
        "word": 20, "normalized_word": 20, "decision": 12, "primary_label": 14,
        "candidate_modes": 22, "confidence": 12, "why_accept": 40,
        "risk_flags": 30, "reject_reason": 35, "review_priority": 14,
        "source_file": 30, "source_line": 12, "pipeline_version": 16,
        "manual_note": 30, "manual_override": 18, "consensus_support": 16,
        "consensus_oppose": 16, "consensus_abstain": 16,
    }
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(col, 16)


def _write_sheet(ws, records: list[dict], cols: list[str], highlight_risk: bool = False):
    """Write header + data rows to a worksheet."""
    ws.append(cols)
    for rec in records:
        row = _flatten_record(rec, cols)
        ws.append(row)
        # Highlight borderline/risky rows
        if highlight_risk:
            flags = rec.get("risk_flags", [])
            priority = _review_priority(rec)
            if flags and priority == "high":
                for cell in ws[ws.max_row]:
                    cell.fill = _RISK_FILL
            elif priority == "medium":
                for cell in ws[ws.max_row]:
                    cell.fill = _BORDER_FILL
    _style_sheet(ws, cols)


# ---------------------------------------------------------------------------
# Main exporter functions
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, saas_records: list[dict], rejected_records: list[dict]):
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    from collections import Counter
    label_dist = Counter(r.get("primary_label", "unknown") for r in saas_records)
    risk_dist = Counter(
        flag for r in saas_records for flag in r.get("risk_flags", [])
    )

    data = [
        ("Total accepted", len(saas_records)),
        ("Total rejected", len(rejected_records)),
        ("Pipeline version", PIPELINE_VERSION),
        ("", ""),
        ("--- Label distribution ---", ""),
    ] + [(f"  label={k}", v) for k, v in sorted(label_dist.items())] + [
        ("", ""),
        ("--- Risk flags ---", ""),
    ] + [(f"  flag={k}", v) for k, v in sorted(risk_dist.items())]

    for row in data:
        ws.append(list(row))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def export_saas_review(saas_records: list[dict]) -> None:
    """Write saas_words_review.xlsx and saas_words_review.csv."""
    accepted = [r for r in saas_records if not r.get("risk_flags")]
    borderline = [r for r in saas_records if r.get("risk_flags")]

    wb = openpyxl.Workbook()

    # Sheet 1: accepted_words
    ws_acc = wb.active
    ws_acc.title = "accepted_words"
    _write_sheet(ws_acc, accepted, ACCEPTED_COLS)

    # Sheet 2: borderline_words
    ws_brd = wb.create_sheet("borderline_words")
    _write_sheet(ws_brd, borderline, ACCEPTED_COLS, highlight_risk=True)

    # Sheet 3: rejected_words placeholder (populated by export_rejected_review)
    ws_rej = wb.create_sheet("rejected_words")
    ws_rej.append(["(see rejected_words_review.xlsx for full reject list)"])

    # Sheet 4: summary
    ws_sum = wb.create_sheet("summary")
    _write_summary_sheet(ws_sum, saas_records, [])

    # Sheet 5: qa_findings (placeholder — populated after QA runs)
    ws_qa = wb.create_sheet("qa_findings")
    ws_qa.append(["(QA findings will be populated after QA run)"])

    OUT_SAAS_REVIEW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_SAAS_REVIEW_XLSX)
    log.info("Wrote %s", OUT_SAAS_REVIEW_XLSX)

    # CSV (accepted + borderline)
    with open(OUT_SAAS_REVIEW_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(ACCEPTED_COLS)
        for rec in saas_records:
            writer.writerow(_flatten_record(rec, ACCEPTED_COLS))
    log.info("Wrote %s", OUT_SAAS_REVIEW_CSV)


def export_rejected_review(rejected_records: list[dict]) -> None:
    """Write rejected_words_review.xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rejected_words"
    _write_sheet(ws, rejected_records, REJECTED_COLS)

    OUT_REJECTED_REVIEW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REJECTED_REVIEW_XLSX)
    log.info("Wrote %s", OUT_REJECTED_REVIEW_XLSX)


def run(saas_records: list[dict], rejected_records: list[dict]) -> None:
    """Step 10: generate all human review files."""
    export_saas_review(saas_records)
    export_rejected_review(rejected_records)
    log.info("Human review export complete.")


# ---------------------------------------------------------------------------
# Streaming variants (memory-efficient)
# ---------------------------------------------------------------------------

def _write_sheet_streaming(ws, source_path: Path, cols: list[str],
                          filter_fn: callable = None, highlight_risk: bool = False):
    """Write worksheet by streaming through source file."""
    # Header
    ws.append(cols)
    _style_header_only(ws)

    # Data rows (streaming)
    with open(source_path, encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rec = json.loads(line)
            if filter_fn and not filter_fn(rec):
                continue
            row = _flatten_record(rec, cols)
            ws.append(row)

            # Highlight borderline/risky rows
            if highlight_risk:
                flags = rec.get("risk_flags", [])
                priority = _review_priority(rec)
                if flags and priority == "high":
                    for cell in ws[ws.max_row]:
                        cell.fill = _RISK_FILL
                elif priority == "medium":
                    for cell in ws[ws.max_row]:
                        cell.fill = _BORDER_FILL

    _style_sheet(ws, cols)


def _style_header_only(ws):
    """Apply header style only (for streaming write where data comes after)."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=False, vertical="center")
    ws.freeze_panes = "A2"


def _write_summary_sheet_streaming(ws, saas_path: Path, rejected_path: Path):
    """Write summary sheet by streaming through files for statistics."""
    ws.append(["Metric", "Value"])
    _style_header_only(ws)

    label_dist = Counter()
    risk_dist = Counter()
    total_accepted = 0

    # Stream saas_words
    with open(saas_path, encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rec = json.loads(line)
            total_accepted += 1
            label_dist[rec.get("primary_label", "unknown")] += 1
            for flag in rec.get("risk_flags", []):
                risk_dist[flag] += 1

    # Count rejected
    total_rejected = 0
    with open(rejected_path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                total_rejected += 1

    data = [
        ("Total accepted", total_accepted),
        ("Total rejected", total_rejected),
        ("Pipeline version", PIPELINE_VERSION),
        ("", ""),
        ("--- Label distribution ---", ""),
    ] + [(f"  label={k}", v) for k, v in sorted(label_dist.items())] + [
        ("", ""),
        ("--- Risk flags ---", ""),
    ] + [(f"  flag={k}", v) for k, v in sorted(risk_dist.items())]

    for row in data:
        ws.append(list(row))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def export_saas_review_streaming() -> None:
    """Write saas_words_review.xlsx and saas_words_review.csv (streaming)."""
    wb = openpyxl.Workbook()

    # Sheet 1: accepted_words (streaming)
    ws_acc = wb.active
    ws_acc.title = "accepted_words"
    _write_sheet_streaming(ws_acc, OUT_SAAS_WORDS, ACCEPTED_COLS,
                          filter_fn=lambda r: not r.get("risk_flags"))

    # Sheet 2: borderline_words (streaming)
    ws_brd = wb.create_sheet("borderline_words")
    _write_sheet_streaming(ws_brd, OUT_SAAS_WORDS, ACCEPTED_COLS,
                          filter_fn=lambda r: bool(r.get("risk_flags")),
                          highlight_risk=True)

    # Sheet 3: rejected_words placeholder
    ws_rej = wb.create_sheet("rejected_words")
    ws_rej.append(["(see rejected_words_review.xlsx for full reject list)"])

    # Sheet 4: summary (streaming)
    ws_sum = wb.create_sheet("summary")
    _write_summary_sheet_streaming(ws_sum, OUT_SAAS_WORDS, OUT_REJECTED_WORDS)

    # Sheet 5: qa_findings placeholder
    ws_qa = wb.create_sheet("qa_findings")
    ws_qa.append(["(QA findings will be populated after QA run)"])

    OUT_SAAS_REVIEW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_SAAS_REVIEW_XLSX)
    log.info("Wrote %s", OUT_SAAS_REVIEW_XLSX)

    # CSV export (streaming)
    with open(OUT_SAAS_REVIEW_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(ACCEPTED_COLS)
        with open(OUT_SAAS_WORDS, encoding="utf-8") as src:
            for line in src:
                if not line.strip():
                    continue
                rec = json.loads(line)
                writer.writerow(_flatten_record(rec, ACCEPTED_COLS))
    log.info("Wrote %s", OUT_SAAS_REVIEW_CSV)


def export_rejected_review_streaming() -> None:
    """Write rejected_words_review.xlsx (streaming)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rejected_words"
    _write_sheet_streaming(ws, OUT_REJECTED_WORDS, REJECTED_COLS)

    OUT_REJECTED_REVIEW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REJECTED_REVIEW_XLSX)
    log.info("Wrote %s", OUT_REJECTED_REVIEW_XLSX)


def run_streaming() -> None:
    """Streaming version: Generate XLSX/CSV by streaming through input files."""
    export_saas_review_streaming()
    export_rejected_review_streaming()
    log.info("Human review export complete (streaming).")
